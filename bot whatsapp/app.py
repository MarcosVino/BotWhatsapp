# Bot de mensagens automaticas
# criado por https://github.com/MarcosVino

# como usar 
# antes de tudo, voce precisa entrar na planilia Mensagens.xlsx e mudar a mensagem e para quem vc vai mandar 
# depois rode o pos.py para pegar as cordenadas corretas da barra de mensagem do wpp e botao de envio e coloque elas no codigo
# e pronto

import openpyxl
import webbrowser
from time import sleep
import pyautogui
import os


# Ler a planilha e guardar as informações
workbook = openpyxl.load_workbook('COLOQUE O CAMINHO DA PLANILIA AQUI')
pagina_msg = workbook['Sheet1']

# roda um laço for para guardar as informações primeiro
for linha in pagina_msg.iter_rows(min_row=2):
    msg_a_mandar = linha[0].value
    telefone = linha[1].value

# abre o link ja com o numero da pessoa que estiver na planilia
link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={telefone}'
webbrowser.open(link_mensagem_whatsapp)
sleep(20)

# aqui que acontece a magica de enviar as mensagens
for linha in pagina_msg.iter_rows(min_row=2):
    msg_a_mandar = linha[0].value
    telefone = linha[1].value
    
    try:

        #aqui voce muda a posição XY de acordo de onde vc quer clicar para selecionar a barra de mensagens para o bot escrever a mensagem no wpp
        pyautogui.click(3000, 996)
        sleep(.2)

        # escreve a msg que esta na planilia
        pyautogui.write(msg_a_mandar, interval=0)
        
        #aqui vc precisa mudar tbm a posição XY de acordo com a posição do botao enviar do seu wpp
        pyautogui.click(3326, 998)
        sleep(.2)
    

    # caso de algum erro, vai imprimir no console e vai guardar as msg com erro + os numeros que nao foram enviados as mensagens no arquivo erros.cvs
    except:
        print(f'Não foi possível enviar mensagem para {msg_a_mandar}')    
        with open('erros.csv','a',newline='',encoding='utf-8') as arquivo:
            arquivo.write(f'{msg_a_mandar},{telefone}{os.linesep}')

    if msg_a_mandar == None:
        break